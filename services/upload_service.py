import csv
import hashlib
import io
from datetime import datetime
from services.db import db
from models import XUUpload, XUUploadDataV2

try:
    from openpyxl import load_workbook  # type: ignore[import]
except ImportError:
    load_workbook = None


def parse_csv_rows(file_bytes, filename):
    if filename.lower().endswith(".xlsx"):
        if load_workbook is None:
            raise RuntimeError("openpyxl kurulmalı: pip install openpyxl")
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if cell is None else str(cell) for cell in row])
        return [row for row in rows if any(cell != "" for cell in row)]

    csv_text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(csv_text.splitlines())
    rows = [row for row in reader if row]
    return rows


def get_next_ticketno():
    last_ticket = db.session.query(db.func.max(XUUpload.ticketno)).scalar()
    if last_ticket is None:
        return 20218
    return last_ticket + 1


def get_next_upload_data_id():
    last_id = db.session.query(db.func.max(XUUploadDataV2.id)).scalar()
    if last_id is None:
        return 1
    return int(last_id) + 1


def compute_sha2_hash(content_bytes):
    return hashlib.sha256(content_bytes).hexdigest()


def process_csv_upload(application_type, upload_type, description, filename, file_bytes, user_ip=None, username=None):
    try:
        rows = parse_csv_rows(file_bytes, filename)
        if len(rows) == 0:
            return {"success": False, "message": "CSV dosyası boş veya geçersiz."}

        row_count = max(0, len(rows) - 1)
        uploaded_at = datetime.utcnow()
        ticketno = get_next_ticketno()
        sha_hash = compute_sha2_hash(file_bytes)

        upload_record = XUUpload(
            ticketno=ticketno,
            applicationcode=application_type[:20],
            filtercode=upload_type[:20],
            outputcode="CSV",
            username=username or "local-user",
            uploadstartdate=uploaded_at,
            uploadenddate=uploaded_at,
            description=description[:50],
            rowcount=row_count,
            filename=filename[:250],
            stampedfilename=filename[:250],
            userip=user_ip or "127.0.0.1",
            sha2hash=sha_hash,
            uploadstatusid=1,
        )
        db.session.add(upload_record)
        db.session.flush()

        upload_data_records = []

        next_id = get_next_upload_data_id()

        for i, row in enumerate(rows):
            if i == 0:
                continue  # header'ı atla

            upload_data_records.append(
                XUUploadDataV2(
                    id=next_id,
                    upload_id=upload_record.id,
                    upload_data=row[0] if len(row) > 0 else "",
                    process_file_name=filename[:300],
                    extra_info1=str(row_count),
                    extra_info2=description[:100],
                    extra_info3=upload_type,
                    extra_info4=application_type,
                )
            )
            next_id += 1

        db.session.add_all(upload_data_records)
        db.session.commit()

        return {
            "success": True,
            "message": "CSV dosyası başarıyla işlendi.",
            "row_count": row_count,
        }
    except Exception as exc:
        db.session.rollback()
        return {"success": False, "message": f"İşlem sırasında hata oluştu: {exc}"}
